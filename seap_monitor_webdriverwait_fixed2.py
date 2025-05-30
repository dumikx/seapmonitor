from selenium import webdriver
from selenium.webdriver.firefox.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import json
import time
import os
import requests
import logging
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

logging.basicConfig(filename='seap_monitor.log', level=logging.INFO, format='%(asctime)s - %(message)s')

CONFIG_PATH = "config.json"
LAST_IDS_PATH = "last_seen_ids.txt"
RESULTS_EXCEL = "seap_results.xlsx"
TELEGRAM_TOKEN = "8156868565:AAFbuMiXP8PnonxKfv-54RXhULxRtNIyRAQ"
TELEGRAM_CHAT_ID = "7111990152"
SEAP_URL = "https://www.e-licitatie.ro/pub/adv-notices/list/1"
DEBUG = True

def init_excel_file():
    if not os.path.exists(RESULTS_EXCEL):
        wb = Workbook()
        ws = wb.active
        ws.title = "Rezultate SEAP"
        headers = ["Data detectării", "Număr anunț", "Titlu", "Cuvânt cheie", "Data limită depunere"]
        ws.append(headers)
        
        # Stilizare header
        for col in range(1, len(headers)+1):
            ws[get_column_letter(col) + '1'].font = Font(bold=True)
        
        wb.save(RESULTS_EXCEL)

def save_to_excel(notice_number, title, keyword, deadline):
    try:
        wb = load_workbook(RESULTS_EXCEL)
        ws = wb.active
        
        # Verificăm dacă anunțul există deja
        for row in ws.iter_rows(min_row=2, max_col=2):
            if row[1].value == notice_number:
                return  # Anunțul există deja, nu îl adăugăm din nou
        
        # Adăugăm noul anunț
        new_row = [
            time.strftime("%Y-%m-%d %H:%M:%S"),
            notice_number,
            title,
            keyword,
            deadline
        ]
        ws.append(new_row)
        wb.save(RESULTS_EXCEL)
        
        if DEBUG:
            print(f"✅ Salvat în Excel: {notice_number} - {title}")
            
    except Exception as e:
        if DEBUG:
            print(f"⚠️ Eroare la salvarea în Excel: {e}")
        logging.error(f"Eroare la salvarea în Excel: {e}")

def load_config():
    with open(CONFIG_PATH, "r", encoding="utf-8") as f:
        return json.load(f)

def load_last_seen_ids():
    if os.path.exists(LAST_IDS_PATH):
        with open(LAST_IDS_PATH, "r") as f:
            return set(line.strip() for line in f.readlines())
    return set()

def save_last_seen_ids(ids):
    existing = load_last_seen_ids()
    all_ids = existing.union(ids)
    with open(LAST_IDS_PATH, "w") as f:
        f.write("\n".join(sorted(all_ids)))

def send_telegram_message(message):
    url = f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/sendMessage"
    data = {"chat_id": TELEGRAM_CHAT_ID, "text": message, "parse_mode": "HTML"}
    requests.post(url, data=data)

def extract_current_notice_ids(driver):
    ids = set()
    blocks = driver.find_elements(By.CSS_SELECTOR, "div.col-md-4")
    for block in blocks:
        if "Numar anunt:" in block.text:
            try:
                ids.add(block.find_element(By.TAG_NAME, "strong").text.strip())
            except:
                continue
    return ids

def process_results(driver, last_seen_ids, current_ids, found_new, ignored_initial=None, source_keyword=None):
    blocks = driver.find_elements(By.CSS_SELECTOR, "div.u-items-list__item")

    for block in blocks:
        try:
            try:
                notice_number_elem = block.find_element(By.CSS_SELECTOR, "div.col-md-4 strong")
                notice_number = notice_number_elem.text.strip()
            except:
                continue

            if "ADV" not in notice_number:
                continue

            if notice_number in last_seen_ids or (ignored_initial and notice_number in ignored_initial):
                if DEBUG:
                    print(f"Ignorat: {notice_number}")
                continue

            try:
                title_elem = block.find_element(By.CSS_SELECTOR, "a.title-entity")
                title = title_elem.text.strip()
            except:
                title = "(fără titlu)"

            found_new[0] = True
            current_ids.add(notice_number)

            deadline_elem = block.find_element(By.XPATH, ".//span[contains(text(),'Data limita depunere oferta')]/strong")
            deadline = deadline_elem.text.strip()

            message = (
                f"🔔 <b>Anunț nou detectat</b>\n"
                f"<b>Număr anunț:</b> {notice_number}\n"
                f"<b>Titlu:</b> {title}\n"
                f"<b>Cuvânt cheie:</b> {source_keyword}\n"
                f"<b>Data limită depunere:</b> {deadline}"
            )
            send_telegram_message(message)
            save_to_excel(notice_number, title, source_keyword, deadline)

            if DEBUG:
                print(f"✅ Notificat: {notice_number} - {title} (găsit cu: '{source_keyword}')")
                logging.info(f"✅ Notificat: {notice_number} - {title} (găsit cu: '{source_keyword}')")

        except Exception as e:
            if DEBUG:
                print(f"⚠️ Eroare la extragere: {e}")
            continue

def perform_search(driver, input_selector, text, wait_time, is_institution=False):
    try:
        input_elem = driver.find_element(By.CSS_SELECTOR, input_selector)
        input_elem.clear()
        input_elem.send_keys(text)
        time.sleep(0.5)

        if is_institution:
        try:
            # Așteaptă ca overlay-ul să dispară
            WebDriverWait(driver, 15).until(
                EC.invisibility_of_element_located((By.CSS_SELECTOR, "div.block-ui-overlay"))
            )

            # Așteaptă apariția dropdown-ului și click pe primul element
            first_option = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.CSS_SELECTOR, "#filterCaDdl_listbox li"))
            )
            first_option.click()
            time.sleep(0.5)
        except Exception as e:
            print(f"⚠️ Nu s-a putut selecta instituția din dropdown: {e}")
            return
                time.sleep(0.5)
            except Exception as e:
                print(f"⚠️ Nu s-a putut selecta instituția din dropdown: {e}")
                return

        input_elem.send_keys(Keys.ENTER)
        try:
            WebDriverWait(driver, 5).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "div.u-items-list__item"))
            )
        except Exception as e:
            print(f"⚠️ Rezultatele nu s-au încărcat la timp pentru: {text} -> {e}")

    except Exception as e:
        print(f"Eroare la căutarea: {text} -> {e}")

def clear_input(driver, input_selector):
    try:
        input_elem = driver.find_element(By.CSS_SELECTOR, input_selector)
        input_elem.clear()
        input_elem.send_keys(Keys.BACKSPACE * 10)
        WebDriverWait(driver, 5).until(
            EC.text_to_be_present_in_element_value((By.CSS_SELECTOR, input_selector), "")
        )
        input_elem.send_keys(Keys.ENTER)
        time.sleep(0.5)
    except Exception as e:
        print(f"Eroare la golirea câmpului: {e}")

def main():
    init_excel_file()
    config = load_config()
    last_seen_ids = load_last_seen_ids()
    current_ids = set()
    found_new = [False]
    institutions = config["institutions"]
    keywords = config["keywords"]

    options = Options()
    options.headless = True
    driver = webdriver.Firefox(options=options)

    try:
        driver.get(SEAP_URL)
        time.sleep(15)

        ignored_initial = extract_current_notice_ids(driver)

        for keyword in keywords:
            perform_search(driver, "input[ng-model='vm.filter.contractObject']", keyword, wait_time=1)
            process_results(driver, last_seen_ids, current_ids, found_new, ignored_initial=ignored_initial, source_keyword=keyword)

        clear_input(driver, "input[ng-model='vm.filter.contractObject']")
        time.sleep(10)

        for institution in institutions:
            perform_search(driver, "input[aria-owns='filterCaDdl_listbox']", institution, wait_time=2, is_institution=True)
            process_results(driver, last_seen_ids, current_ids, found_new, ignored_initial=ignored_initial, source_keyword=institution)

        if not found_new[0]:
            send_telegram_message("📭 Niciun anunț nou găsit după filtrele setate.")
        save_last_seen_ids(current_ids)

    finally:
        driver.quit()

if __name__ == "__main__":
    main()